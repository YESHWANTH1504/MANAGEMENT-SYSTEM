import io
import datetime
from typing import Optional
from fastapi import APIRouter, Depends, HTTPException, status, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from app.db.session import get_db
from app.db import models
from app.api import deps
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors

router = APIRouter()

@router.get("/reports/excel")
def export_reports_excel(
    intern_id: Optional[int] = Query(None),
    start_date: Optional[datetime.date] = Query(None),
    end_date: Optional[datetime.date] = Query(None),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(deps.get_current_active_mentor_or_admin)
):
    """
    Admins & Mentors: Export daily activity reports to an Excel sheet.
    """
    query = db.query(models.DailyReport).join(models.User, models.DailyReport.user_id == models.User.id)
    
    if intern_id:
        query = query.filter(models.DailyReport.user_id == intern_id)
        
    if start_date:
        query = query.filter(models.DailyReport.date >= start_date)
    if end_date:
        query = query.filter(models.DailyReport.date <= end_date)
        
    reports = query.order_by(models.DailyReport.date.asc()).all()

    # Create Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Daily Activity Reports"

    # Setup Styling
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")

    # Define Header
    headers = [
        "Intern Name", "Internship ID", "Date", "Task Title", 
        "Description", "Hours Worked", "Technologies Used", 
        "Challenges Faced", "Learning Outcomes", "Tomorrow's Plan", "Status"
    ]
    ws.append(headers)

    # Style Header Row
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    # Add Data
    for report in reports:
        # Load Intern profile
        intern = db.query(models.Intern).filter(models.Intern.user_id == report.user_id).first()
        intern_name = intern.full_name if intern else "System"
        internship_id = intern.internship_id if intern else "N/A"
        
        row = [
            intern_name,
            internship_id,
            report.date.strftime("%Y-%m-%d"),
            report.task_title,
            report.description,
            report.hours_worked,
            report.technologies_used or "",
            report.challenges_faced or "",
            report.learning_outcomes or "",
            report.tomorrow_plan or "",
            report.status.upper()
        ]
        ws.append(row)

    # Adjust Column Widths
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = max(max_len + 3, 10)

    # Save to Bytes stream
    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    filename = f"IMMS_Reports_Export_{datetime.date.today()}.xlsx"
    return StreamingResponse(
        file_stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@router.get("/attendance/excel")
def export_attendance_excel(
    intern_id: Optional[int] = Query(None),
    start_date: Optional[datetime.date] = Query(None),
    end_date: Optional[datetime.date] = Query(None),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(deps.get_current_active_mentor_or_admin)
):
    """
    Admins & Mentors: Export Attendance sheets to Excel.
    """
    query = db.query(models.Attendance)
    
    if intern_id:
        query = query.filter(models.Attendance.user_id == intern_id)
        
    if start_date:
        query = query.filter(models.Attendance.date >= start_date)
    if end_date:
        query = query.filter(models.Attendance.date <= end_date)
        
    attendance_records = query.order_by(models.Attendance.date.asc()).all()

    # Create Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance Logs"

    # Setup Styling
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="375623", end_color="375623", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")

    headers = ["Intern Name", "Internship ID", "Date", "Check-In", "Check-Out", "Status", "IP Address"]
    ws.append(headers)

    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    for record in attendance_records:
        intern = db.query(models.Intern).filter(models.Intern.user_id == record.user_id).first()
        intern_name = intern.full_name if intern else "System"
        internship_id = intern.internship_id if intern else "N/A"
        
        check_in = record.check_in_time.strftime("%I:%M %p") if record.check_in_time else "N/A"
        check_out = record.check_out_time.strftime("%I:%M %p") if record.check_out_time else "N/A"
        
        row = [
            intern_name,
            internship_id,
            record.date.strftime("%Y-%m-%d"),
            check_in,
            check_out,
            record.status.upper(),
            record.ip_address or "N/A"
        ]
        ws.append(row)

    # Adjust Column Widths
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = max(max_len + 3, 10)

    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    filename = f"IMMS_Attendance_Export_{datetime.date.today()}.xlsx"
    return StreamingResponse(
        file_stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@router.get("/portfolio/pdf/{intern_id}")
def export_portfolio_pdf(
    intern_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(deps.get_current_user)
):
    """
    Generate and download a comprehensive progress portfolio PDF for an intern or employee.
    Users can download their own; Admins/Mentors can download any.
    """
    intern = db.query(models.Intern).filter(models.Intern.id == intern_id).first()
    is_employee = False
    
    if not intern:
        emp = db.query(models.Employee).filter(models.Employee.id == intern_id).first()
        if not emp:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Profile not found")
        
        class MappedProfile:
            def __init__(self, e):
                self.id = e.id
                self.user_id = e.user_id
                self.full_name = e.full_name
                self.internship_id = e.employee_id
                self.internship_domain = e.designation
                self.department = e.department
                self.college_name = "N/A"
                self.start_date = e.joining_date
                self.end_date = "N/A"
                self.project_name = e.project_name
                self.programming_languages = e.programming_languages
                self.frameworks = e.frameworks
                self.tools_used = e.tools_used
                self.databases_used = e.databases_used
        intern = MappedProfile(emp)
        is_employee = True
        
    # Access checks
    if current_user.role in ["intern", "employee"] and current_user.id != intern.user_id:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Access denied")

    # Fetch related records
    reports = db.query(models.DailyReport).filter(
        models.DailyReport.user_id == intern.user_id,
        models.DailyReport.status == "approved"
    ).order_by(models.DailyReport.date.asc()).all()

    attendance_stats = db.query(models.Attendance).filter(models.Attendance.user_id == intern.user_id).all()
    present_days = sum(1 for a in attendance_stats if a.status in ["present", "late"])
    total_days = len(attendance_stats)
    att_rate = round((present_days / total_days) * 100, 1) if total_days > 0 else 100.0

    # Build PDF
    pdf_buffer = io.BytesIO()
    doc = SimpleDocTemplate(pdf_buffer, pagesize=letter, rightMargin=40, leftMargin=40, topMargin=40, bottomMargin=40)
    
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=24,
        leading=28,
        textColor=colors.HexColor('#1A365D'),
        spaceAfter=15
    )
    
    subtitle_style = ParagraphStyle(
        'DocSubtitle',
        parent=styles['Heading3'],
        fontSize=12,
        leading=16,
        textColor=colors.HexColor('#4A5568'),
        spaceAfter=25
    )

    body_style = styles['Normal']
    bold_body = ParagraphStyle('BoldBody', parent=body_style, fontName='Helvetica-Bold')

    elements = []
    
    # Document Header
    elements.append(Paragraph("Internship Performance Portfolio", title_style))
    elements.append(Paragraph(f"Internship Management & Monitoring System (IMMS) - Generated {datetime.date.today().strftime('%B %d, %Y')}", subtitle_style))
    elements.append(Spacer(1, 10))

    # Profile Section
    profile_data = [
        [Paragraph("Intern Name:", bold_body), Paragraph(intern.full_name, body_style), Paragraph("Internship ID:", bold_body), Paragraph(intern.internship_id, body_style)],
        [Paragraph("Domain:", bold_body), Paragraph(intern.internship_domain or "N/A", body_style), Paragraph("Department:", bold_body), Paragraph(intern.department or "N/A", body_style)],
        [Paragraph("College:", bold_body), Paragraph(intern.college_name or "N/A", body_style), Paragraph("Duration:", bold_body), Paragraph(f"{intern.start_date or 'N/A'} to {intern.end_date or 'N/A'}", body_style)],
        [Paragraph("Project Name:", bold_body), Paragraph(intern.project_name or "N/A", body_style), Paragraph("Attendance Rate:", bold_body), Paragraph(f"{att_rate}% ({present_days}/{total_days} days)", body_style)]
    ]
    
    profile_table = Table(profile_data, colWidths=[100, 160, 100, 160])
    profile_table.setStyle(TableStyle([
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#CBD5E1')),
        ('BACKGROUND', (0,0), (0,-1), colors.HexColor('#F8FAFC')),
        ('BACKGROUND', (2,0), (2,-1), colors.HexColor('#F8FAFC')),
        ('PADDING', (0,0), (-1,-1), 6),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
    ]))
    elements.append(profile_table)
    elements.append(Spacer(1, 20))

    # Technology Section
    elements.append(Paragraph("Technologies Applied & Learnt", styles['Heading2']))
    techs_summary = f"<b>Languages:</b> {intern.programming_languages or 'N/A'}<br/><b>Frameworks:</b> {intern.frameworks or 'N/A'}<br/><b>Tools/DBs:</b> {intern.tools_used or 'N/A'} / {intern.databases_used or 'N/A'}"
    elements.append(Paragraph(techs_summary, styles['BodyText']))
    elements.append(Spacer(1, 20))

    # Work History Section
    elements.append(Paragraph("Approved Daily Reports & Tasks Log", styles['Heading2']))
    
    table_headers = [Paragraph("<b>Date</b>", body_style), Paragraph("<b>Task Title</b>", body_style), Paragraph("<b>Hours</b>", body_style), Paragraph("<b>Tech Stack</b>", body_style)]
    report_table_data = [table_headers]
    
    for rep in reports:
        report_table_data.append([
            Paragraph(rep.date.strftime("%Y-%m-%d"), body_style),
            Paragraph(rep.task_title, body_style),
            Paragraph(str(rep.hours_worked), body_style),
            Paragraph(rep.technologies_used or "N/A", body_style)
        ])
        
    if len(reports) == 0:
        report_table_data.append([Paragraph("No approved reports found.", body_style), "", "", ""])

    report_table = Table(report_table_data, colWidths=[80, 240, 50, 150])
    report_table.setStyle(TableStyle([
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#CBD5E1')),
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#E2E8F0')),
        ('PADDING', (0,0), (-1,-1), 6),
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
    ]))
    
    elements.append(report_table)

    # Build Doc
    doc.build(elements)
    pdf_buffer.seek(0)

    filename = f"IMMS_Intern_Portfolio_{intern.internship_id}.pdf"
    return StreamingResponse(
        pdf_buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
